import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from timeis import timeis, green, yellow, line, white, tic, toc


class ReadXlsx:

    def __init__(self, filename, sheets_name):
        print(f"{timeis()} {green}opening xlsx")
        self.workbook = load_workbook(filename)
        self.soup = BeautifulSoup(features='xml')  # We use BeautifulSoup for XML handling
        self.sheets_name = sheets_name
        self.df = {}
        
        if isinstance(sheets_name, str):
            self.sheets_name = [sheets_name]
            
        for sheet_name in sheets_name:
            sheet = self.workbook[sheet_name]
            sheet_rows = list(sheet.iter_rows(values_only=True))
            header = sheet_rows[0]
            data = [dict(zip(header, row)) for row in sheet_rows[1:]]
        
            self.df[sheet_name] = pd.DataFrame(data)
            self.apply_bold_styles(sheet, sheet_name)

    def apply_bold_styles(self, sheet, sheet_name):
        """Apply bold formatting tags where appropriate."""
        rows = list(sheet.iter_rows())
        num_rows = self.df[sheet_name].shape[0]
        num_columns = self.df[sheet_name].shape[1]

        for i, row in enumerate(rows):
            if i >= num_rows:
                break  # Avoid going out of bounds in the DataFrame
            for j, cell in enumerate(row):
                if j >= num_columns:
                    continue  # Avoid going out of bounds in the DataFrame
                if cell.font and cell.font.bold:
                    if isinstance(self.df[sheet_name].iloc[i, j], str):
                        self.df[sheet_name].iloc[i, j] = f"<b>{self.df[sheet_name].iloc[i, j]}</b>"


    def process_and_save_csv(self):
        print(f"{timeis()} {green}processing for anki")
            
        test1 = self.df['analysis']['#'] == "1"
        test2 = self.df['analysis']['meaning'] != ""
        filter = test1 & test2
        self.df['analysis'] = self.df['analysis'][filter]
        self.df['analysis'].drop(["#", "x", "comments"], axis = 1, inplace=True)
        self.df['analysis'].drop(self.df['analysis'].iloc[:, 20:], axis = 1, inplace=True)
        self.df['analysis']['feedback'] = (
            f"Spot a mistake? <a class=\"link\" href=\"https://docs.google.com/forms/d/e/1FAIpQLSdG6zKDtlwibtrX-cbKVn4WmIs8miH4VnuJvb7f94plCDKJyA/viewform?usp=pp_url&entry.438735500="
            + self.df['analysis'].pali_1 + "&entry.1433863141=Anki\">Fix it here</a>."
        )

        rows = self.df['analysis'].shape[0]
        columns = self.df['analysis'].shape[1]
        self.df['analysis'].to_csv('Pātimokkha Word by Word.csv', sep='\t', index=False, header=True, quoting=1)

        print(f"{timeis()} {green}saving {white}{rows} {green}rows {white}{columns} {green}columns")
        # print(self.df['analysis'])

    def process_and_save_column_names(self):
        print(f"{timeis()} {green}saving column names to text file")

        with open('../../sasanarakkha/study-tools/anki-style/field-list-pat.txt', 'w', encoding='utf-8') as txt_file:
            txt_file.write("\n".join(self.df['analysis'].columns) + "\n")

        print(f"{timeis()} {green}column names saved to Column_Names.txt")


if __name__ == '__main__':
    tic()
    print(f"{timeis()} {line}")
    print(f"{timeis()} {yellow}converting xlsx to csv")
    print(f"{timeis()} {line}")
    a = ReadXlsx("original_sources/Pātimokkha Word by Word.xlsx", ['analysis'])
    a.process_and_save_csv()
    a.process_and_save_column_names()
    toc()
